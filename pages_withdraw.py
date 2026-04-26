"""pages_withdraw.py — Lab Parfumo PO Pro
เบิกสินค้าออกจากสต๊อก + ดูประวัติการเบิก"""
from datetime import date, datetime, timedelta
import streamlit as st

import database as db
from helpers import current_user, is_admin, uid, uname, fmt_date


def render_withdraw():
    """หน้าหลัก: เบิกสินค้า + ประวัติ"""
    st.markdown("""
    <div class="page-title-block">
        <div class="page-title-text">เบิกสินค้า</div>
        <div class="page-title-sub">บันทึกการใช้สินค้าจากสต็อก + ดูประวัติ</div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["➕ เบิกใหม่", "📜 ประวัติการเบิก"])

    with tab1:
        _render_withdraw_form()

    with tab2:
        _render_withdraw_history()


# ==================================================================
# Tab 1: เบิกใหม่
# ==================================================================

def _render_withdraw_form():
    """ฟอร์มเบิกสินค้า"""
    eq_list = db.get_equipment_list(active_only=True)

    if not eq_list:
        st.info("ยังไม่มีสินค้าใน Catalog")
        return

    # ===== Filter + Search + Sort =====
    col1, col2, col3 = st.columns([3, 2, 2])
    with col1:
        search = st.text_input(
            "🔍 ค้นหาสินค้า",
            placeholder="ชื่อสินค้า / SKU",
            key="withdraw_search",
        ).strip().lower()
    with col2:
        cats = ["ทั้งหมด"] + db.get_categories()
        selected_cat = st.selectbox("📂 หมวด", cats, key="withdraw_cat")
    with col3:
        sort_by = st.selectbox(
            "🔃 จัดเรียง",
            ["ชื่อ A→Z", "ชื่อ Z→A",
             "สต็อกมาก→น้อย", "สต็อกน้อย→มาก",
             "หมวด"],
            key="withdraw_sort",
        )

    # Filter
    filtered = eq_list
    if search:
        filtered = [
            e for e in filtered
            if search in (e.get('name') or '').lower()
            or search in (e.get('sku') or '').lower()
        ]
    if selected_cat != "ทั้งหมด":
        filtered = [e for e in filtered if e.get('category') == selected_cat]

    # ซ่อนสินค้าที่สต็อก = 0
    show_zero = st.checkbox("แสดงสินค้าที่หมดด้วย", value=False,
                              key="withdraw_show_zero")
    if not show_zero:
        filtered = [e for e in filtered if (e.get('stock') or 0) > 0]

    # Sort
    if sort_by == "ชื่อ A→Z":
        filtered = sorted(filtered, key=lambda e: (e.get('name') or '').lower())
    elif sort_by == "ชื่อ Z→A":
        filtered = sorted(filtered, key=lambda e: (e.get('name') or '').lower(), reverse=True)
    elif sort_by == "สต็อกมาก→น้อย":
        filtered = sorted(filtered, key=lambda e: float(e.get('stock') or 0), reverse=True)
    elif sort_by == "สต็อกน้อย→มาก":
        filtered = sorted(filtered, key=lambda e: float(e.get('stock') or 0))
    elif sort_by == "หมวด":
        filtered = sorted(filtered, key=lambda e: ((e.get('category') or 'zzz').lower(),
                                                       (e.get('name') or '').lower()))

    st.caption(f"พบ **{len(filtered)}** รายการที่เบิกได้")

    if not filtered:
        st.warning("ไม่พบสินค้าตามที่ค้นหา")
        return

    # ===== แสดงเป็น 3 columns =====
    for row_start in range(0, len(filtered), 3):
        row = filtered[row_start:row_start + 3]
        cols = st.columns(3)
        for i, eq in enumerate(row):
            with cols[i]:
                _render_withdraw_card(eq)


def _render_withdraw_card(eq):
    """การ์ดสินค้าพร้อมปุ่มเบิก"""
    eq_id = eq['id']
    name = eq.get('name', '-')
    sku = eq.get('sku') or '-'
    stock = float(eq.get('stock') or 0)
    unit = eq.get('unit', 'ชิ้น')

    # รูป
    images = list(eq.get('image_urls') or [])
    if eq.get('image_url') and eq['image_url'] not in images:
        images.insert(0, eq['image_url'])

    with st.container(border=True):
        # รูป
        if images:
            st.markdown(
                f'<div style="width:100%; aspect-ratio:1/1; '
                f'background:#F4F6FA; border-radius:8px; overflow:hidden; '
                f'margin-bottom:8px;">'
                f'<img src="{images[0]}" '
                f'style="width:100%; height:100%; object-fit:cover;"/>'
                f'</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="width:100%; aspect-ratio:1/1; '
                'background:#F4F6FA; border-radius:8px; '
                'display:flex; align-items:center; justify-content:center; '
                'font-size:48px; margin-bottom:8px;">🧴</div>',
                unsafe_allow_html=True,
            )

        # ชื่อ
        st.markdown(
            f'<div style="font-weight:500; font-size:15px; '
            f'white-space:nowrap; overflow:hidden; text-overflow:ellipsis;" '
            f'title="{name}">{name}</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<div style="font-size:12px; color:#888;">'
            f'SKU: {sku} | 📂 {eq.get("category", "-")}</div>',
            unsafe_allow_html=True,
        )

        # Stock indicator
        if stock == 0:
            stock_color, stock_emoji = "#A32D2D", "🔴"
            stock_label = "หมด"
        elif stock < 10:
            stock_color, stock_emoji = "#BA7517", "🟡"
            stock_label = f"เหลือ {stock:,.0f}"
        else:
            stock_color, stock_emoji = "#1D9E75", "🟢"
            stock_label = f"คงเหลือ {stock:,.0f}"

        st.markdown(
            f'<div style="margin:8px 0; padding:6px 0; '
            f'border-top:1px solid #E5E7EB; '
            f'color:{stock_color}; font-weight:500; font-size:13px;">'
            f'{stock_emoji} {stock_label} {unit}</div>',
            unsafe_allow_html=True,
        )

        # ===== ปุ่มเปิด/ปิดฟอร์ม =====
        form_key = f'_show_form_{eq_id}'
        is_open = st.session_state.get(form_key, False)

        if stock == 0:
            st.button("❌ สต็อกหมด", use_container_width=True,
                       disabled=True, key=f'btn_zero_{eq_id}')
        elif is_open:
            # แสดงฟอร์ม
            with st.form(f'wf_{eq_id}', clear_on_submit=False):
                c1, c2 = st.columns(2)
                with c1:
                    qty = st.number_input(
                        "จำนวน",
                        min_value=1,
                        max_value=int(stock),
                        value=1,
                        step=1,
                        key=f'qty_{eq_id}',
                    )
                with c2:
                    w_date = st.date_input(
                        "วันที่ใช้",
                        value=date.today(),
                        key=f'date_{eq_id}',
                    )

                purpose = st.text_input(
                    "ใช้ทำอะไร / ใช้ที่ไหน",
                    placeholder="เช่น ผลิต CELEB lot 24, ตัวอย่างลูกค้า",
                    key=f'purpose_{eq_id}',
                )

                bc1, bc2 = st.columns(2)
                with bc1:
                    submit = st.form_submit_button(
                        "✅ บันทึกเบิก",
                        type="primary",
                        use_container_width=True,
                    )
                with bc2:
                    cancel = st.form_submit_button(
                        "ยกเลิก",
                        use_container_width=True,
                    )

                if submit:
                    if not purpose.strip():
                        st.error("กรุณากรอกใช้ทำอะไร")
                    else:
                        # แปลง date → datetime (บังคับเที่ยงวันเพื่อหลีกเลี่ยง timezone)
                        w_dt = datetime.combine(w_date, datetime.min.time())
                        result = db.create_withdrawal(
                            equipment_id=eq_id,
                            qty=qty,
                            purpose=purpose.strip(),
                            withdrawn_by=uid(),
                            withdrawn_by_name=uname(),
                            withdrawn_at=w_dt,
                        )
                        if result:
                            st.session_state[form_key] = False
                            st.success(f"✅ เบิก {qty} {unit} แล้ว")
                            st.rerun()
                if cancel:
                    st.session_state[form_key] = False
                    st.rerun()
        else:
            if st.button("📤 เบิกสินค้า", use_container_width=True,
                          type="primary", key=f'open_{eq_id}'):
                st.session_state[form_key] = True
                st.rerun()


# ==================================================================
# Tab 2: ประวัติการเบิก
# ==================================================================

def _render_withdraw_history():
    """ประวัติการเบิก + filter + search + sort"""
    user = current_user()

    # ===== Row 1: Scope + Period + Equipment filter =====
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        # admin เห็นของทุกคน, staff เห็นของตัวเอง (default)
        if is_admin():
            scope = st.selectbox(
                "👤 ดูของ",
                ["ของฉัน", "ทุกคน"],
                key="hist_scope",
            )
        else:
            scope = "ของฉัน"
            st.caption("👤 ดูของฉัน")
    with fc2:
        period = st.selectbox(
            "📅 ช่วงเวลา",
            ["7 วันล่าสุด", "30 วันล่าสุด", "90 วันล่าสุด", "ทั้งหมด"],
            index=1,
            key="hist_period",
        )
    with fc3:
        eq_list = db.get_equipment_list(active_only=False)
        eq_options = ["ทั้งหมด"] + [f"{e['name']}" for e in eq_list]
        eq_filter = st.selectbox(
            "📦 สินค้า",
            eq_options,
            key="hist_eq",
        )

    # ===== Row 2: Search + Sort =====
    sc1, sc2 = st.columns([3, 2])
    with sc1:
        search = st.text_input(
            "🔍 ค้นหา",
            placeholder="ชื่อสินค้า / วัตถุประสงค์ / ผู้เบิก",
            key="hist_search",
        ).strip().lower()
    with sc2:
        sort_by = st.selectbox(
            "🔃 จัดเรียง",
            ["วันที่ใหม่→เก่า", "วันที่เก่า→ใหม่",
             "จำนวนมาก→น้อย", "จำนวนน้อย→มาก",
             "ชื่อสินค้า A→Z"],
            key="hist_sort",
        )

    # ===== Build query =====
    user_id_filter = None if (is_admin() and scope == "ทุกคน") else uid()

    start_date = None
    if period == "7 วันล่าสุด":
        start_date = datetime.now() - timedelta(days=7)
    elif period == "30 วันล่าสุด":
        start_date = datetime.now() - timedelta(days=30)
    elif period == "90 วันล่าสุด":
        start_date = datetime.now() - timedelta(days=90)

    eq_id_filter = None
    if eq_filter != "ทั้งหมด":
        match = next((e for e in eq_list if e['name'] == eq_filter), None)
        if match:
            eq_id_filter = match['id']

    withdrawals = db.get_withdrawals(
        equipment_id=eq_id_filter,
        user_id=user_id_filter,
        start_date=start_date,
    )

    # ===== Apply search =====
    if search:
        withdrawals = [
            w for w in withdrawals
            if search in (w.get('equipment_name') or '').lower()
            or search in (w.get('purpose') or '').lower()
            or search in (w.get('withdrawn_by_name') or '').lower()
            or search in (w.get('notes') or '').lower()
        ]

    # ===== Apply sort =====
    def _date_key(w):
        """key สำหรับเรียงตามวันที่ — fallback เป็น created_at ถ้าวันที่เท่ากัน"""
        try:
            primary = datetime.fromisoformat(
                (w.get('withdrawn_at') or '').replace('Z', '+00:00')
            )
        except Exception:
            primary = datetime.min
        try:
            secondary = datetime.fromisoformat(
                (w.get('created_at') or '').replace('Z', '+00:00')
            )
        except Exception:
            secondary = datetime.min
        return (primary, secondary)

    if sort_by == "วันที่ใหม่→เก่า":
        withdrawals = sorted(withdrawals, key=_date_key, reverse=True)
    elif sort_by == "วันที่เก่า→ใหม่":
        withdrawals = sorted(withdrawals, key=_date_key)
    elif sort_by == "จำนวนมาก→น้อย":
        withdrawals = sorted(withdrawals, key=lambda w: float(w.get('qty', 0) or 0), reverse=True)
    elif sort_by == "จำนวนน้อย→มาก":
        withdrawals = sorted(withdrawals, key=lambda w: float(w.get('qty', 0) or 0))
    elif sort_by == "ชื่อสินค้า A→Z":
        withdrawals = sorted(withdrawals, key=lambda w: (w.get('equipment_name') or '').lower())

    # ===== Summary =====
    if withdrawals:
        total_items = len(withdrawals)
        total_qty = sum(float(w.get('qty', 0) or 0) for w in withdrawals)
        unique_eq = len(set(w.get('equipment_id') for w in withdrawals))

        m1, m2, m3 = st.columns(3)
        with m1:
            st.metric("📊 จำนวนครั้ง", f"{total_items:,}")
        with m2:
            st.metric("📦 รวมจำนวน", f"{total_qty:,.0f}")
        with m3:
            st.metric("🏷️ สินค้าที่เบิก", f"{unique_eq}")

    st.markdown("---")

    # ===== Header: count + export buttons =====
    hc1, hc2, hc3 = st.columns([3, 1, 1])
    with hc1:
        st.caption(f"พบ **{len(withdrawals)}** รายการ")
    with hc2:
        if withdrawals:
            csv_bytes = _build_csv(withdrawals)
            now_str = datetime.now().strftime('%Y%m%d_%H%M')
            st.download_button(
                "📥 ดาวน์โหลด CSV",
                data=csv_bytes,
                file_name=f"withdrawals_{now_str}.csv",
                mime="text/csv",
                use_container_width=True,
                key="dl_csv_w",
            )
    with hc3:
        if withdrawals:
            xlsx_bytes = _build_xlsx(withdrawals)
            if xlsx_bytes:
                now_str = datetime.now().strftime('%Y%m%d_%H%M')
                st.download_button(
                    "📊 Excel",
                    data=xlsx_bytes,
                    file_name=f"withdrawals_{now_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_xlsx_w",
                )

    if not withdrawals:
        st.info("ยังไม่มีประวัติการเบิกในช่วงเวลานี้")
        return

    # ===== Table =====
    for w in withdrawals:
        with st.container(border=True):
            c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 2, 1])
            with c1:
                st.markdown(f"**📦 {w.get('equipment_name', '-')}**")
                purpose = w.get('purpose', '')
                if purpose:
                    st.caption(f"📝 {purpose}")
            with c2:
                qty = float(w.get('qty', 0) or 0)
                unit = w.get('unit', 'ชิ้น')
                st.markdown(
                    f'<div style="color:#4A6FA5; font-weight:500;">'
                    f'➖ {qty:,.0f} {unit}</div>',
                    unsafe_allow_html=True,
                )
            with c3:
                wd_date = w.get('withdrawn_at')
                if wd_date:
                    try:
                        dt = datetime.fromisoformat(wd_date.replace('Z', '+00:00'))
                        st.caption(f"📅 {dt.strftime('%d/%m/%Y')}")
                    except Exception:
                        st.caption(f"📅 {wd_date[:10]}")
            with c4:
                st.caption(f"👤 {w.get('withdrawn_by_name', '-')}")
            with c5:
                # admin ลบได้
                if is_admin():
                    del_key = f'del_w_{w["id"]}'
                    if st.session_state.get(del_key):
                        if st.button("⚠️", key=f'cd_{w["id"]}',
                                      help="ยืนยันลบ + คืนสต็อก",
                                      use_container_width=True):
                            if db.delete_withdrawal(w['id'], restore_stock=True):
                                st.session_state.pop(del_key, None)
                                st.success("ลบ + คืนสต็อกแล้ว")
                                st.rerun()
                    else:
                        if st.button("🗑️", key=f'd_{w["id"]}',
                                      help="ลบ",
                                      use_container_width=True):
                            st.session_state[del_key] = True
                            st.rerun()


# ==================================================================
# Export Helpers
# ==================================================================

def _build_csv(withdrawals):
    """สร้าง CSV bytes (รองรับภาษาไทย — UTF-8 with BOM ให้ Excel เปิดถูก)"""
    import io
    import csv

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        'วันที่ใช้',
        'สินค้า',
        'SKU',
        'หมวด',
        'จำนวน',
        'หน่วย',
        'ใช้ทำอะไร',
        'ผู้เบิก',
        'หมายเหตุ',
        'บันทึกเมื่อ',
    ])

    # ดึง equipment สำหรับเอา SKU + หมวด
    try:
        eq_list = db.get_equipment_list(active_only=False)
        eq_map = {e['id']: e for e in eq_list}
    except Exception:
        eq_map = {}

    for w in withdrawals:
        eq = eq_map.get(w.get('equipment_id'), {})
        # วันที่ใช้
        wd = w.get('withdrawn_at', '')
        try:
            wd_dt = datetime.fromisoformat(wd.replace('Z', '+00:00'))
            wd_str = wd_dt.strftime('%Y-%m-%d')
        except Exception:
            wd_str = wd[:10] if wd else ''

        # บันทึกเมื่อ
        ca = w.get('created_at', '')
        try:
            ca_dt = datetime.fromisoformat(ca.replace('Z', '+00:00'))
            ca_str = ca_dt.strftime('%Y-%m-%d %H:%M')
        except Exception:
            ca_str = ca[:16] if ca else ''

        writer.writerow([
            wd_str,
            w.get('equipment_name', ''),
            eq.get('sku') or '-',
            eq.get('category') or '-',
            float(w.get('qty', 0) or 0),
            w.get('unit', ''),
            w.get('purpose', ''),
            w.get('withdrawn_by_name', ''),
            w.get('notes', ''),
            ca_str,
        ])

    # เพิ่ม BOM (\ufeff) ให้ Excel รู้ว่าเป็น UTF-8 → เปิดภาษาไทยถูกต้อง
    csv_str = '\ufeff' + output.getvalue()
    return csv_str.encode('utf-8')


def _build_xlsx(withdrawals):
    """สร้าง Excel bytes — return None ถ้า openpyxl ไม่มี"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "ประวัติการเบิก"

    headers = [
        'วันที่ใช้', 'สินค้า', 'SKU', 'หมวด',
        'จำนวน', 'หน่วย', 'ใช้ทำอะไร', 'ผู้เบิก',
        'หมายเหตุ', 'บันทึกเมื่อ',
    ]

    # เขียน header + style
    header_fill = PatternFill(start_color="4A6FA5", end_color="4A6FA5", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    # ดึง equipment
    try:
        eq_list = db.get_equipment_list(active_only=False)
        eq_map = {e['id']: e for e in eq_list}
    except Exception:
        eq_map = {}

    # เขียน data
    for row_idx, w in enumerate(withdrawals, 2):
        eq = eq_map.get(w.get('equipment_id'), {})

        wd = w.get('withdrawn_at', '')
        try:
            wd_dt = datetime.fromisoformat(wd.replace('Z', '+00:00'))
            wd_str = wd_dt.strftime('%Y-%m-%d')
        except Exception:
            wd_str = wd[:10] if wd else ''

        ca = w.get('created_at', '')
        try:
            ca_dt = datetime.fromisoformat(ca.replace('Z', '+00:00'))
            ca_str = ca_dt.strftime('%Y-%m-%d %H:%M')
        except Exception:
            ca_str = ca[:16] if ca else ''

        row_data = [
            wd_str,
            w.get('equipment_name', ''),
            eq.get('sku') or '-',
            eq.get('category') or '-',
            float(w.get('qty', 0) or 0),
            w.get('unit', ''),
            w.get('purpose', ''),
            w.get('withdrawn_by_name', ''),
            w.get('notes', ''),
            ca_str,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border_thin
            cell.alignment = Alignment(vertical='center')
            if col_idx == 5:  # คอลัมน์จำนวน
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # ตั้งความกว้าง column
    widths = [12, 28, 12, 16, 10, 10, 30, 16, 24, 18]
    for i, w_val in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w_val

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

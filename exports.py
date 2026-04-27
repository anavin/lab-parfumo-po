"""exports.py - shared CSV / Excel builders ใช้ทั้งระบบ

Pattern: ทุก builder คืน `bytes` พร้อมใส่ใน st.download_button ได้เลย
- CSV จะนำหน้าด้วย BOM ﻿ ให้ Excel เปิดภาษาไทยได้ถูกต้อง
- Excel ใช้ openpyxl — ถ้าไม่มีให้ return None (ผู้เรียก fallback ไป CSV)
"""
import io
import csv
from datetime import datetime


# ==================================================================
# Internal — Excel styling (re-use across exports)
# ==================================================================
def _excel_setup(ws_title="Sheet1"):
    """สร้าง Workbook + style header — return (wb, ws, styles dict) หรือ None"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = ws_title[:31]  # Excel limit

    styles = {
        'header_fill': PatternFill(start_color="4A6FA5",
                                     end_color="4A6FA5",
                                     fill_type="solid"),
        'header_font': Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        'border': Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        ),
        'center': Alignment(horizontal='center', vertical='center'),
        'right': Alignment(horizontal='right', vertical='center'),
        'left_top': Alignment(vertical='center', wrap_text=True),
    }
    return wb, ws, styles


def _excel_write_headers(ws, headers, styles):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = styles['center']
        cell.border = styles['border']


def _excel_save(wb):
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _col_letter(i):
    """1 -> A, 26 -> Z, 27 -> AA"""
    letters = ""
    while i > 0:
        i, rem = divmod(i - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _excel_set_widths(ws, widths):
    for i, w_val in enumerate(widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w_val


def _fmt_iso_date(s):
    """ISO datetime string → 'YYYY-MM-DD' หรือคืน '' ถ้าแปลงไม่ได้"""
    if not s:
        return ''
    try:
        dt = datetime.fromisoformat(str(s).replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return str(s)[:10]


def _fmt_iso_dt(s):
    """ISO datetime string → 'YYYY-MM-DD HH:MM' หรือคืน '' ถ้าแปลงไม่ได้"""
    if not s:
        return ''
    try:
        dt = datetime.fromisoformat(str(s).replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d %H:%M')
    except Exception:
        return str(s)[:16]


# ==================================================================
# PO List Export
# ==================================================================
PO_HEADERS = [
    'PO Number', 'สถานะ', 'วันที่สร้าง',
    'ผู้สั่ง', 'Supplier',
    'รายการ', 'จำนวน items',
    'ยอดก่อน VAT', 'ส่วนลด', 'ค่าส่ง', 'VAT', 'ยอดสุทธิ',
    'วันที่สั่ง', 'คาดได้รับ', 'ได้รับจริง',
    'หมายเหตุ',
]


def _po_row(po):
    items = po.get('items') or []
    items_str = ", ".join(
        f"{it.get('name', '')} × {int(it.get('qty', 0) or 0):,}"
        for it in items
    )
    return [
        po.get('po_number', ''),
        po.get('status', ''),
        _fmt_iso_date(po.get('created_at')),
        po.get('created_by_name', ''),
        po.get('supplier_name') or '',
        items_str,
        len(items),
        float(po.get('subtotal', 0) or 0),
        float(po.get('discount', 0) or 0),
        float(po.get('shipping_fee', 0) or 0),
        float(po.get('vat', 0) or 0),
        float(po.get('total', 0) or 0),
        _fmt_iso_date(po.get('ordered_date')),
        _fmt_iso_date(po.get('expected_date')),
        _fmt_iso_date(po.get('received_date')),
        po.get('notes', ''),
    ]


def po_list_to_csv(pos):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(PO_HEADERS)
    for po in pos:
        writer.writerow(_po_row(po))
    return ('﻿' + output.getvalue()).encode('utf-8')


def po_list_to_xlsx(pos):
    setup = _excel_setup("รายการ PO")
    if setup is None:
        return None
    wb, ws, styles = setup
    _excel_write_headers(ws, PO_HEADERS, styles)
    for row_idx, po in enumerate(_po_row(p) for p in pos):
        for col_idx, val in enumerate(po, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=val)
            cell.border = styles['border']
            # Right-align ตัวเลข (column 7 = จำนวน items, 8-12 = เงิน)
            if col_idx in (7, 8, 9, 10, 11, 12):
                cell.alignment = styles['right']
                if col_idx in (8, 9, 10, 11, 12):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = styles['left_top']
    _excel_set_widths(ws, [14, 16, 12, 18, 22, 38, 8,
                           14, 12, 12, 12, 14,
                           12, 12, 12, 30])
    ws.freeze_panes = "A2"
    return _excel_save(wb)


# ==================================================================
# Equipment Catalog Export
# ==================================================================
EQ_HEADERS = [
    'SKU', 'ชื่อสินค้า', 'หมวด', 'หน่วย',
    'สต็อก', 'Reorder Level',
    'ราคาต้นทุนล่าสุด', 'รายละเอียด',
    'Approval', 'สร้างเมื่อ',
]


def _eq_row(eq):
    return [
        eq.get('sku') or '',
        eq.get('name', ''),
        eq.get('category') or '',
        eq.get('unit', 'ชิ้น'),
        int(eq.get('stock') or 0),
        int(eq.get('reorder_level') or 0),
        float(eq.get('last_cost') or 0),
        (eq.get('description') or '').replace('\n', ' ')[:500],
        eq.get('approval_status') or 'approved',
        _fmt_iso_date(eq.get('created_at')),
    ]


def equipment_to_csv(equipment_list):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EQ_HEADERS)
    for eq in equipment_list:
        writer.writerow(_eq_row(eq))
    return ('﻿' + output.getvalue()).encode('utf-8')


def equipment_to_xlsx(equipment_list):
    setup = _excel_setup("Catalog")
    if setup is None:
        return None
    wb, ws, styles = setup
    _excel_write_headers(ws, EQ_HEADERS, styles)
    for row_idx, eq in enumerate(_eq_row(e) for e in equipment_list):
        for col_idx, val in enumerate(eq, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=val)
            cell.border = styles['border']
            if col_idx in (5, 6, 7):
                cell.alignment = styles['right']
                if col_idx == 7:
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = styles['left_top']
    _excel_set_widths(ws, [14, 28, 16, 8, 10, 12, 16, 36, 12, 12])
    ws.freeze_panes = "A2"
    return _excel_save(wb)


# ==================================================================
# Helper — Streamlit download buttons (CSV + Excel pair)
# ==================================================================
def render_download_pair(prefix_label, csv_bytes, xlsx_bytes, filename_base,
                         key_prefix=""):
    """Render 2 columns of download buttons — CSV + Excel

    ใช้กับ st.columns() ก่อน เช่น:
        c1, c2 = st.columns(2)
        with c1: st.download_button(... csv_bytes ...)
        with c2: st.download_button(... xlsx_bytes ...)

    หรือเรียก helper นี้แทน:
        render_download_pair("📥", csv, xlsx, "po_list_2026")
    """
    import streamlit as st
    now_str = datetime.now().strftime('%Y%m%d_%H%M')
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            f"{prefix_label} CSV",
            data=csv_bytes,
            file_name=f"{filename_base}_{now_str}.csv",
            mime="text/csv",
            use_container_width=True,
            key=f"{key_prefix}_csv",
        )
    with c2:
        if xlsx_bytes:
            st.download_button(
                f"{prefix_label} Excel",
                data=xlsx_bytes,
                file_name=f"{filename_base}_{now_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"{key_prefix}_xlsx",
            )
        else:
            st.button(f"{prefix_label} Excel (ต้องติดตั้ง openpyxl)",
                      use_container_width=True, disabled=True,
                      key=f"{key_prefix}_xlsx_disabled")
